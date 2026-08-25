#!/usr/bin/env python3
"""
Generate Apify inputs for the one thing the 2026-08 backfill did not collect:
review text.

That run used `maxReviews: 0`, so it filled ratings, review *counts* and photos
and left the review bodies behind. The result is visible in two places:

- "<venue name> reviews" queries earn impressions at positions 5-8 (474 in the
  2026-06..08 GSC export) against pages that contain no review text at all.
- app/activities/[niche]/[slug]/page.tsx gates its second ad on
  `hasReviewText`, because a page of structured facts with no prose is the
  shape AdSense reads as made-for-advertising. 1,767 spa pages are in that
  state today.

Ratings and photos are done — 2,251 of 2,311 spa venues qualify, only 52 lack
a rating and 4 lack a photo — so this run asks for reviews and nothing else,
which is also the cheapest way to ask.

Usage:
    python scripts/make_review_backfill_inputs.py            # all niches, 500/chunk
    python scripts/make_review_backfill_inputs.py --niche spa --chunk 600
    python scripts/make_review_backfill_inputs.py --reviews 3
"""
from __future__ import annotations

import argparse
import json
import urllib.parse
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
NICHE_DIR = ROOT / "data" / "by-niche"
OUT_DIR = ROOT.parent / "apify_inputs" / "reviews_backfill"

NICHES = ["spa", "yoga-pilates", "coworking", "diving", "cooking", "muay-thai", "wellness"]


def gsc_impressions(xlsx: str) -> dict[str, float]:
    """
    Impressions per URL slug, from a Search Console Pages export.

    A run costs money per venue, so which venues go first is a real decision.
    Trust Score is a guess at what Google will reward; this is a record of what
    Google is already showing. A venue whose page is being served to people —
    and that ranks at 5-8 for "<name> reviews" with no review text on it — is
    where review text converts to clicks fastest.

    Slug is the key because the same venue appears under /activities/{niche}/
    and /{lang}/place/, and both sets of impressions belong to it.
    """
    try:
        import openpyxl
    except ImportError:
        print("  --gsc needs openpyxl (pip install openpyxl); ordering by Trust Score instead")
        return {}
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    if "Pages" not in wb.sheetnames:
        print(f"  {xlsx} has no 'Pages' sheet; ordering by Trust Score instead")
        return {}
    out: dict[str, float] = {}
    for row in wb["Pages"].iter_rows(min_row=2, values_only=True):
        url, _clicks, impr = str(row[0] or ""), row[1], row[2] or 0
        slug = urllib.parse.unquote(url.rstrip("/").rsplit("/", 1)[-1])
        if slug:
            out[slug] = out.get(slug, 0) + float(impr)
    return out


def needs_reviews(p: dict) -> bool:
    if p.get("permanently_closed"):
        return False
    # A venue with no rating was never on Google properly; reviews won't be
    # there either. Only ask for what can actually come back.
    if not p.get("review_count"):
        return False
    return not p.get("reviews_sample")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--niche", action="append", help="repeatable; default all")
    ap.add_argument("--chunk", type=int, default=500, help="place IDs per input file")
    ap.add_argument("--reviews", type=int, default=5, help="reviews per venue")
    ap.add_argument(
        "--gsc",
        help="GSC Pages export (.xlsx). Venues whose page already earns "
             "impressions are written to the first chunk.",
    )
    ap.add_argument(
        "--top",
        type=int,
        help="Emit a single file holding the best N venues across every niche, "
             "instead of chunking per niche. For a run with a fixed budget.",
    )
    args = ap.parse_args()

    niches = args.niche or NICHES
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    impressions = gsc_impressions(args.gsc) if args.gsc else {}
    total = 0
    written: list[str] = []
    pool: list[dict] = []

    for niche in niches:
        src = NICHE_DIR / f"{niche}.json"
        if not src.exists():
            print(f"  skip {niche}: no dataset")
            continue
        places = json.loads(src.read_text(encoding="utf-8"))["places"]
        wanted = [
            p for p in places
            if needs_reviews(p) and str(p.get("id", "")).startswith("ChIJ")
        ]
        # Proven demand first, then Trust Score as the fallback ordering. The
        # dataset's own order carries no meaning, and with a budget that covers
        # one chunk, chunk 1 is the only one that may get run.
        wanted.sort(
            key=lambda p: (
                -impressions.get(p.get("slug", ""), 0.0),
                -(p.get("trust_score") or 0),
            )
        )
        ids = [p["id"] for p in wanted]
        if not ids:
            print(f"  {niche}: nothing missing")
            continue
        seen = sum(1 for p in wanted if impressions.get(p.get("slug", ""), 0) > 0)
        note = f" ({seen} already earning impressions)" if impressions else ""
        print(f"  {niche}: {len(ids)} venues need review text{note}")
        total += len(ids)

        if args.top:
            pool.extend(wanted)
            continue

        for i in range(0, len(ids), args.chunk):
            part = ids[i : i + args.chunk]
            n = i // args.chunk + 1
            name = f"{niche}_reviews_{n}.json" if len(ids) > args.chunk else f"{niche}_reviews.json"
            payload = {
                "language": "en",
                "scrapePlaceDetailPage": True,
                # Everything except reviews is already in the dataset; asking
                # again would cost more and overwrite good data with the same
                # data. enrich_from_apify.py only fills empty fields anyway,
                # except photos, which it deliberately refreshes — so leaving
                # maxImages at 0 also keeps this run from touching them.
                "maxImages": 0,
                "maxReviews": args.reviews,
                "reviewsSort": "newest",
                "maxQuestions": 0,
                "scrapeContacts": False,
                "maximumLeadsEnrichmentRecords": 0,
                "skipClosedPlaces": False,
                "placeIds": part,
            }
            out = OUT_DIR / name
            out.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
            written.append(f"{name} ({len(part)})")

    if args.top:
        pool.sort(
            key=lambda p: (
                -impressions.get(p.get("slug", ""), 0.0),
                -(p.get("trust_score") or 0),
            )
        )
        picked = pool[: args.top]
        payload = {
            "language": "en",
            "scrapePlaceDetailPage": True,
            "maxImages": 0,
            "maxReviews": args.reviews,
            "reviewsSort": "newest",
            "maxQuestions": 0,
            "scrapeContacts": False,
            "maximumLeadsEnrichmentRecords": 0,
            "skipClosedPlaces": False,
            "placeIds": [p["id"] for p in picked],
        }
        out = OUT_DIR / f"priority_{args.top}.json"
        out.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        by_niche: dict[str, int] = {}
        for p in picked:
            by_niche[p.get("niche", "?")] = by_niche.get(p.get("niche", "?"), 0) + 1
        proven = sum(1 for p in picked if impressions.get(p.get("slug", ""), 0) > 0)
        print(f"\n{out.name}: {len(picked)} of {total} venues")
        print(f"  {proven} already earn impressions; the rest are the highest Trust Scores")
        print("  " + ", ".join(f"{k} {v}" for k, v in sorted(by_niche.items(), key=lambda kv: -kv[1])))
        return 0

    print(f"\n{total} venues across {len(written)} input files in {OUT_DIR}")
    for w in written:
        print(f"  {w}")
    print(
        "\nRun the smallest file first, then merge it with:\n"
        "  python scripts/enrich_from_apify.py <export.json> --dry-run\n"
        "and check that reviews_sample is being filled before running the rest."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
